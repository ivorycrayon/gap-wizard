#Ver 1.3 - Updated 7/22/22

import os
import sys
import openpyxl as ox #needs pip install openpyxl
import time
import operator
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment

class Associate:
    def __init__(self, login):
        self.login = login
        self.shift_start = ""
        self.shift_end = ""
        self.end_time = ""
        self.current_time = ""
        self.time_spent = 0
        self.shift_time_total = 0
        self.shift_break_deduction = 0
        self.tot_total = 0
        self.tot_list = []
        self.old_row = []
        self.row_list = []
        self.row_len = 0
        self.shift_utilization = 0

    def __str__(self) -> str:
        return(f"""
Associate: {self.login}
Start Time: {self.shift_start}
End time: {self.shift_end}
Total Time (in hours): {self.shift_time_total}
""")

class MainWindow(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack(fill=tk.BOTH, expand=True)
        self.create_widgets()

    def create_widgets(self):
        self.frame8 = tk.Frame(root)
        self.frame10 = tk.Frame(root)

        self.frame8.pack(fill=tk.BOTH, expand=True, padx = 1, pady = 1)
        self.frame10.pack(fill=tk.BOTH, expand=True, padx = 1, pady = 1)

        self.label8 = tk.Label(
            master = self.frame8,
            text = "Gap Threshold (minutes):",
            width = 25,
            borderwidth = 1, 
            relief = "solid",
            anchor = 'w',
            font = 'Calibri 12 bold'
        )

        self.entry8 = tk.Entry(master = self.frame8, width = 5)
        self.entry8.insert(tk.END, "10")

        self.label8.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.entry8.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.button1 = tk.Button(
            master = self.frame10,
            command = self.on_button,
            text = "Run Wizard",
            width = 20,
            height = 5,
            bg = "blue",
            fg = "yellow",
            font = 'Calibri 12 bold'
        )
        self.button1.pack()

        self.button2 = tk.Button(
            master = self.frame10,
            command = self.on_quit_button,
            text = "Quit",
            width = 10,
            height = 2,
            bg = "red",
            fg = "yellow",
            font = 'Calibri 12 bold'
        )
        self.button2.pack()
        

    def on_button(self):
        threshold = float(self.entry8.get())
        run_genie(threshold)
        
    def on_quit_button(self):
        root.destroy()

def set_border(ws, cell_range, style):
    rows = ws[cell_range]
    side = Side(border_style=style, color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

def run_genie(threshold):
    associates = []
    timeformat = "%m/%d/%Y %I:%M:%S %p"

    gap_file = 'gap.xlsx'

    t0 = time.time()

    try:
        transactions = ox.load_workbook(filename = gap_file, read_only = False)
    except OSError:
        tk.messagebox.showerror("File Load Error", "Please ensure there is a file named \'gap.xlsx\' in the same folder as this program")
        
    sheet = transactions['Sheet1']

    for row in sheet.iter_rows(min_row = 1, max_row = 1):
        i = 0
        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for cell in row:
            if cell.value == "Start Date":
                start_time_index = i
                start_time_index_alpha = alphabet[i]
            if cell.value == "End Date":
                end_time_index = i
                end_time_index_alpha = alphabet[i]
            if cell.value == "Tran Type":
                tran_type_index = i
                tran_type_index_alpha = alphabet[i]
            if cell.value == "Location Id":
                loc_id_index = i
                loc_id_index_alpha = alphabet[i]
            if cell.value == "Location Id 2":
                loc_id_2_index = i
                loc_id_2_index_alpha = alphabet[i]
            if cell.value == "Hu Id 2":
                hu_id_2_index = i
                hu_id_2_index_alpha = alphabet[i]
            if cell.value == "Tran Qty":
                tran_qty_index = i
                tran_qty_index_alpha = alphabet[i]
            if cell.value == "Employee Id":
                employee_id_index = i
                employee_id_index_alpha = alphabet[i]
            if cell.value == "Item Number":
                item_number_index = i
                item_number_index_alpha = alphabet[i]
            i += 1
    
    t1 = time.time()
    print(f"Time to load workbook: {t1 - t0} seconds")
    t0 = time.time()

    #build list of associate objects
    for cell in sheet[employee_id_index_alpha]:
        if cell.value not in (associate.login for associate in associates) and cell.coordinate != employee_id_index_alpha + '1':
                associates.append(Associate(cell.value))

    old_row = []

    for row in sheet.iter_rows(min_row = 2):
        associndex = [x.login for x in associates].index(row[employee_id_index].value)

        #this fixes the 12:00:00 AM inconsistency in the dates appearing as simply the date and no time
        modifydate = "" 
        if(len(row[start_time_index].value) < 16):
            modifydate += row[start_time_index].value
            modifydate += " 12:00:00 AM"
            row[start_time_index].value = modifydate
            modifydate = "" 
        if(len(row[end_time_index].value) < 16):
            modifydate += row[end_time_index].value
            modifydate += " 12:00:00 AM"
            row[end_time_index].value = modifydate
            modifydate = "" 

        if not associates[associndex].old_row: #check empty
            for cell in row:
                associates[associndex].old_row.append(cell.value)

        if(associates[associndex].end_time == ""):
            associates[associndex].end_time = datetime.strptime(row[end_time_index].value, timeformat)
            associates[associndex].shift_end = datetime.strptime(row[end_time_index].value, timeformat)
        if(associates[associndex].current_time == ""):
            associates[associndex].current_time = datetime.strptime(row[start_time_index].value, timeformat)

        associates[associndex].current_time = datetime.strptime(row[end_time_index].value, timeformat)
        associates[associndex].shift_start = datetime.strptime(row[end_time_index].value, timeformat)

        associates[associndex].time_spent = (associates[associndex].end_time - associates[associndex].current_time).total_seconds()/60.0

        if(associates[associndex].time_spent > threshold):
            associates[associndex].tot_total += associates[associndex].time_spent
            associates[associndex].tot_list.append(associates[associndex].time_spent)

            new_row = []
            for cell in row:
                new_row.append(cell.value)

            associates[associndex].row_list.append(associates[associndex].old_row)
            associates[associndex].row_list.append(new_row)
            associates[associndex].old_row = new_row
        else:
            new_row = []
            for cell in row:
                new_row.append(cell.value)

            associates[associndex].old_row = new_row

        associates[associndex].end_time = datetime.strptime(row[end_time_index].value, timeformat)

    t1 = time.time()
    print(f"Time to iterate the sheet and do stuff: {t1 - t0} seconds")

    template = 'GAPS.xlsx'

    try:
        os.chdir(".\\Templates\\")
    except:
        tk.messagebox.showerror("Folder Missing Error", "Please ensure there is a folder named \'Templates\' in the same folder as this program")
    
    try:
        Gaps = ox.load_workbook(template)
    except:
        tk.messagebox.showerror("File Missing Error", "Please ensure there is a file named \'GAPS.xlsx\' in the Templates folder")    

    sheet = Gaps["Gaps"]
    sheet2 = Gaps["Totals"]

    associates.sort(key = operator.attrgetter('tot_total'), reverse = True)

    for associate in associates:
        associate.shift_time_total = (associate.shift_end - associate.shift_start).total_seconds()/3600.0 #gives time in hrs
        if(associate.shift_time_total > 6.5):
            associate.shift_break_deduction = 70.0 #lunch & 2 breaks
        elif(associate.shift_time_total > 4.5):
            associate.shift_break_deduction = 55.0 #lunch & break
        elif(associate.shift_time_total > 2.5):
            associate.shift_break_deduction = 15.0 #one break

    j = 2
    ch = 'A'
    i = 1
    for associate in associates:
        print(associate)
        associate.row_len = len(associate.row_list)
        first_cell = True
        x = 0
        for row in associate.row_list:
            for column in row:
                cell = ch + str(j)
                if(ch != 'A'):
                    sheet[cell] = column
                if(ch == 'A' and first_cell):
                    sheet[cell] = "Total TOT for " + associate.login + ": " + str(round(associate.tot_total - associate.shift_break_deduction)) + " minutes (Breaks excluded)"
                    first_cell = False
                    sheet[cell].font = Font(bold = True)
                if(i % 2 == 0 and ch == 'A'):
                    sheet[cell] = str(round(associate.tot_list[x],1)) + " minute gap."
                    x += 1

                ch = chr(ord(ch) + 1)
            if(i % 2 == 0):
                j += 1
            i += 1
            ch = 'A'
            j += 1
        j += 1

    # a = 2
    # t = 3
    # ch1 = 'A'
    # ch2 = 'T'
    # i = 1
    # for row in sheet.iter_rows(min_row = 2):
    #     if(i % 2 == 0):
    #         cell1 = ch1 + str(a)
    #         cell2 = ch2 + str(t)
    #         set_border(sheet, cell1 + ":" + cell2, "thin")
    #         a += 2
    #         t += 2
    #     i += 1

    a = 2
    t = 1
    ch1 = 'A'
    ch2 = 'T'
    for associate in associates:
        t += associate.row_len + int((associate.row_len / 2)) - 1
        cell1 = ch1 + str(a)
        cell2 = ch2 + str(t)
        set_border(sheet, cell1 + ":" + cell2, "thick")
        a += associate.row_len + 2 + int((associate.row_len / 2)) - 1
        t += 2

    #begin Totals page
    
    tot_total_list = []

    for associate in associates:
        tot_total_list.append(associate.tot_total)
        if(associate.shift_time_total > 0):
            associate.shift_utilization = 100 * ((associate.shift_time_total * 60.0) - (associate.tot_total - associate.shift_break_deduction)) / (associate.shift_time_total * 60.0)
        else: 
            associate.shift_utilization = 0
    
    i = 2

    associates.sort(key = operator.attrgetter('shift_utilization'), reverse = True)

    for associate in associates:
        ch = 'A'
        cell = ch + str(i)
        sheet2[cell] = associate.login
        #set_border(sheet, cell, "thin")

        ch = 'B'
        cell = ch + str(i)
        if associate.tot_total - associate.shift_break_deduction < 60:
            sheet2[cell] = str(round(associate.tot_total - associate.shift_break_deduction)) + " minutes"
        else:
            sheet2[cell] = str(round((associate.tot_total - associate.shift_break_deduction) / 60.0, 1)) + " hours"

        ch = 'C'
        cell = ch + str(i)  
        sheet2[cell] = str(round(associate.shift_time_total, 1)) + " hours"

        ch = 'D'
        cell = ch + str(i)  
        sheet2[cell] = int(round(associate.shift_utilization)) 
        sheet2[cell].alignment = Alignment(horizontal = "center")


        i += 1

    a = 2
    d = 2
    ch1 = 'A'
    ch2 = 'D'
    for associate in associates:
        cell1 = ch1 + str(a)
        cell2 = 'A' + str(d)
        set_border(sheet2, cell1 + ":" + cell2, "thin")
        cell2 = 'B' + str(d)
        set_border(sheet2, cell1 + ":" + cell2, "thin")
        cell2 = 'C' + str(d)
        set_border(sheet2, cell1 + ":" + cell2, "thin")
        cell2 = 'D' + str(d)
        set_border(sheet2, cell1 + ":" + cell2, "thin")
        a += 1
        d += 1

            
    os.chdir("..")
    if not os.path.exists(".\\Gaps\\"):
        os.makedirs(".\\Gaps\\")
    os.chdir(".\\Gaps\\")

    now = datetime.now().strftime("%m-%d-%y %H;%M;%S")
    
    file_out = "TIME GAPS " + now + ".xlsx"
    if(os.path.exists(file_out)):
        os.remove(file_out)
    Gaps.save(file_out)

    tk.messagebox.showinfo("Success!~", "File successfully generated in the Gaps folder.")
    os.chdir("..")

    #for associate in associates:
    #    print(associate)

root = tk.Tk()
root.title("Time Gap Genie")
root.geometry("300x330")
try:
    os.chdir(".\\Templates\\")
    root.iconbitmap("icon.ico")
except:
    tk.messagebox.showinfo("File Missing Error", "Missing icon.ico file from Templates folder.")
os.chdir("..")

Window = MainWindow(master = root)
Window.mainloop()
